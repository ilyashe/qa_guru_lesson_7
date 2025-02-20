import csv
import zipfile
from pypdf import PdfReader
from openpyxl import load_workbook
from tests.conftest import ZIP_FILE


def test_csv_file():
    with zipfile.ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('file_example_CSV.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            seventh_row = csvreader[6]

            assert seventh_row[0] == '6'
            assert seventh_row[1] == 'Gaston'


def test_pdf_file():
    with zipfile.ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('file_example_PDF.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)
            assert number_of_pages == 297

            page = reader.pages[2]
            text = page.extract_text()
            assert 'По моя начин' in text


def test_xlsx_file():
    with zipfile.ZipFile(ZIP_FILE) as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as csv_file:
            workbook = load_workbook(csv_file)
            sheet = workbook.active
            assert sheet.cell(row=10, column=3).value == 'Weiland'
            assert sheet.max_row == 51
            assert sheet.max_column == 8
